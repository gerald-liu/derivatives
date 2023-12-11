import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from matplotlib import pyplot as plt

from product.params import Params
from product.option import SharkFin
from pricer.pricer import get_pricing_results
from io_product.surface import Surface

# key: variable name
# value: column names in the table
DICT_RESULTS = {
    'price': 'Price',
    'delta': 'Delta',
    'gamma': 'Gamma',
    'vega': 'Vega'
}

DICT_SHARKFIN = {
    'name': 'Product',
    'T': 'Tenor (m)',
    'X': 'Strike',
    'H': 'KO Barrier',
    'freq': 'KO Frequency',
    'r_min': 'Min Return',
    'r_ko': 'KO Return',
    'pt': 'Participation Rate',
    'S': 'Spot',
    'sigma': 'Vol',
    'r': 'r',
    'coc': 'coc'
}

MAP_PRODUCT_DICTS = {
    'sharkfin': DICT_SHARKFIN
}

# A4 dimension in inches
PAPER_SIZE = (8.27, 11.69)

class XlsxIO():
    def __init__(
        self, product_type, year_len = 365,
        skiptop = 1, skipleft = 0, skipfooter = 0
    ):
        self._product_type = product_type
        self._year_len = year_len
        self._skiptop = skiptop
        self._skipleft = skipleft
        self._skipfooter = skipfooter

        self._in_filename = None
        self._out_filename = None
        self._out_filename_img = None
        self._img_filenames = []

        self._dict_product = MAP_PRODUCT_DICTS[self._product_type]
        self._df_in = pd.DataFrame()
        self._df_out = pd.DataFrame()
        self._list_products = []
        
    def read(self, filename, sheetname):
        self._in_filename = filename
        self._df_in = pd.read_excel(
            io = self._in_filename,
            sheet_name = sheetname,
            names = list(self._dict_product.values()),
            usecols = range(self._skipleft, len(self._dict_product)),
            skiprows = range(0, self._skiptop),
            skipfooter = self._skipfooter
        )
    
    def get_dt(self, freq):
        if freq[-1] == 'D':
            return float(freq[:-1]) / self._year_len
        elif freq[-1] == 'M':
            return float(freq[:-1]) / 12
        elif freq[-1] == 'Y':
            return float(freq[:-1])
        else:
            raise Exception('Error: invalid KO Frequency')

    def init_sharkfin(self):
        list_o = []
        for _, row in self._df_in.iterrows():
            name = str(row[DICT_SHARKFIN['name']]).lower()
            if 'call' in name:
                is_call = True
            elif 'put' in name:
                is_call = False
            else:
                raise Exception('Error: invalid Product name')

            S = float(row[DICT_SHARKFIN['S']])
            T = float(row[DICT_SHARKFIN['T']]) / 12
            X = float(row[DICT_SHARKFIN['X']])
            H = float(row[DICT_SHARKFIN['H']])
            r_min = float(row[DICT_SHARKFIN['r_min']])
            r_ko = float(row[DICT_SHARKFIN['r_ko']])
            pt = float(row[DICT_SHARKFIN['pt']])

            params = Params(
                float(row[DICT_SHARKFIN['sigma']]),
                float(row[DICT_SHARKFIN['r']]),
                float(row[DICT_SHARKFIN['coc']])
            )

            dt = self.get_dt(str(row[DICT_SHARKFIN['freq']]))
            
            sf = SharkFin(S, T, params, X, H, dt, r_min, r_ko, pt, is_call)
            list_o.append(sf)
        
        return list_o

    def init_products(self):
        self._list_products = [] # reset
        if self._product_type == 'sharkfin':
            self._list_products += self.init_sharkfin()

    def get_products(self, refresh=False):
        if refresh or not len(self._list_products):
            self.init_products()
        return self._list_products
    
    # method = {'closed-form', 'monte-carlo'}
    # 'monte-carlo' args = [dt]
    def get_results(self, method = 'closed-form', inplace = True, price_only = False):
        if price_only:
            df = pd.DataFrame(columns=[DICT_RESULTS['price']])
        else:
            df = pd.DataFrame(columns=list(DICT_RESULTS.values()))

        for p in self.get_products():
            if method == 'closed-form':
                args = []
            elif method == 'monte-carlo':
                args = [p.dt]
            else:
                raise Exception('Error: Pricing method undefined.')
        
            df.loc[len(df.index)] = get_pricing_results(
                p, method = method, args = args, price_only = price_only
            )

        if inplace:
            self._df_out = df
        
        return df

    # n_cols: # of subplots per row (# of cols)
    # figsize[0]: width of A4
    def plot(self, out_filename, S_endp, T_endp, n_cols = 2, N = 100, price_only = False):
        products = self.get_products()
        if price_only:
            tasks = ['price']
        else:
            tasks = list(DICT_RESULTS.keys())

        n_rows = len(products) * len(tasks) // 2

        fig, axs = plt.subplots(
            nrows = n_rows, ncols = n_cols,
            figsize = (PAPER_SIZE[0], PAPER_SIZE[0]/2 * n_rows),
            layout = 'constrained', subplot_kw = {"projection": "3d"}
        )
        fig.set_constrained_layout_pads(w_pad = 0.2, h_pad = 0.2)

        plot_id = 0
        plot_row = 0
        plot_col = 0
        
        S_range = np.linspace(S_endp[0], S_endp[1], N)

        # T: time to maturity
        if hasattr(products[0], '_dt'): # discretely monitored
            dt = products[0].dt
            T_first = np.ceil(T_endp[1] / dt)
            T_last = np.floor(T_endp[0] / dt)
            T_num = T_first - T_last + 1 # max # of observations
            T_size = min(T_num, N)
            T_range = dt * np.linspace(T_last, T_first, T_size, dtype=int)
        else:
            T_range = np.linspace(T_endp[0], T_endp[1], N)

        for p in products:
            surface = Surface(S_range, T_range, p)
            surface.generate()

            for task in list(DICT_RESULTS.keys()):
                surface.plot(axs[plot_row][plot_col], task)
                
                plot_id += 1
                plot_row, plot_col = divmod(plot_id, 2)

        plt.savefig(out_filename)
        plt.close(fig)
    
    def write_results(self, out_filename, sheetname):
        # df_print = self._df_out.map(lambda x: '{:.2%}'.format(x))
        # with pd.ExcelWriter(out_filename, mode='a', if_sheet_exists='overlay') as writer:
        #     df_print.to_excel(
        #         writer, sheet_name = sheetname,
        #         header = True, index = False,
        #         startrow = self._skiptop,
        #         startcol = self._skipleft + self._df_in.shape[1]
        #     )

        self._out_filename = out_filename
        if self._in_filename is None:
            raise Exception('Error: Input file undefined.')

        wb = load_workbook(self._in_filename)
        ws = wb[sheetname]

        # 1-based
        start_row = self._skiptop + 1
        start_col = self._skipleft + self._df_in.shape[1] + 1

        for j in range(self._df_out.shape[1]):
            cell = ws.cell(row = start_row, column = start_col + j)
            cell.value = self._df_out.columns[j]
            cell.fill = PatternFill(fill_type='solid', fgColor='1F4E78')
            cell.font = Font(b=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for i in range(self._df_out.shape[0]):
            for j in range(self._df_out.shape[1]):
                cell = ws.cell(row = start_row + 1 + i, column = start_col + j)
                cell.value = self._df_out.iloc[i, j]
                cell.number_format = '0.00%'
                cell.fill = PatternFill(fill_type='solid', fgColor='FFFF00')
                cell.font = Font(color='FF0000')
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        wb.save(out_filename)

    def write_image(self, img_filename, xlsx_filename, sheetname):
        if self._out_filename is None:
            raise Exception('Error: Output file with data is undefined.')
        
        wb = load_workbook(self._out_filename)
        ws = wb[sheetname]
        img = Image(img_filename)

        spacing = 2

        # 1-based
        img_row = self._skiptop + self._df_in.shape[0] + 1 + spacing
        img_col = self._skipleft + 1

        img.anchor = f'{get_column_letter(img_col)}{img_row + 1}'
        ws.add_image(img)
        wb.save(xlsx_filename)
    