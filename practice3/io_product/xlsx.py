import numpy as np
import pandas as pd
import xlwings as xw
from openpyxl.utils.cell import get_column_letter

from matplotlib import pyplot as plt

from product.constants import *
from product.params import Params
from product.option import init_option
from product.portfolio import Portfolio
from pricer.pricer import get_pricing_results
from io_product.surface import Surface


DICT_PORTFOLIO = {
    'weight': '权重'
}

DICT_OPTION = {
    'product': '产品结构',
    'ticker': '挂钩标的',
    'T': '期限',
    'r_min': '最低收益',
    'r_max': '最高收益',
    'X_lo': '低行权价',
    'X_hi': '高行权价',
    'H_lo': '下障碍价',
    'H_hi': '上障碍价',
    'K': '敲出收益',
    'pt': '参与率',
    'sigma': 'Sigma',
    # 'r': 'r',
    'b': 'coc'
}

DICT_RESULTS = {
    'price': 'Price (%)',
    'delta': 'Delta (%)',
    'gamma': 'Gamma (%)',
    'vega': 'Vega (%)'
}

# input file specs
SKIPROWS = []
SKIPLEFT = 0
SKIPFOOTER = 1

PAPER_SIZE = (8.27, 11.69) # A4 size in inches


""" HARDCODED
The original input table leaves sigma and coc empty.
Fill in values for reading.
"""
def amend_input(filename):
    with xw.Book(filename) as wb:
        sheet = wb.sheets[0]
        sheet.range('M2:M19').value = 0.3
        sheet.range('N2:N19').value = 0.05
        wb.save()


class XlsxIO():
    def __init__(self, dict_portfolio=DICT_PORTFOLIO, dict_option=DICT_OPTION,
        skipleft=SKIPLEFT, skiprows=SKIPROWS, skipfooter=SKIPFOOTER
    ):
        self._dict_portfolio = dict_portfolio
        self._dict_option = dict_option
        self._skipleft = skipleft
        self._skiprows = skiprows
        self._skipfooter = skipfooter

        self._df_in = pd.DataFrame()
        self._df_out_products = pd.DataFrame()
        
        self._portfolio = Portfolio()

        self._surface = None
        self._out_dfs = {}

    def read(self, filename, sheet_idx=0):
        self._df_in = pd.read_excel(
            io = filename,
            sheet_name = sheet_idx,
            # names = list(self._dict_cols.keys()),
            usecols = range(self._skipleft, len(self._dict_portfolio) + len(self._dict_option)),
            skiprows = self._skiprows,
            skipfooter = self._skipfooter
        )
        self._df_in = self._df_in.fillna(0)

    def init_portfolio(self):
        for _, row in self._df_in.iterrows():          
            option_args = {}
            for key, value in self._dict_option.items():
                option_args[key] = row[value]

            option_args['T'] = option_args['T'] / 12 # month to year

            # HARD-CODED PART
            option_args['r'] = 0 # missing column in input table
            option_args.pop('ticker', None) # unused

            option = init_option(**option_args)
            weight = row[self._dict_portfolio['weight']]

            self._portfolio.add_product(option, weight)

    def get_portfolio(self, refresh=False):
        if refresh:
            self._portfolio.reset()
        if self._portfolio.is_empty():
            self.init_portfolio()
        return self._portfolio
    
    # method = {'closed-form', 'monte-carlo'}
    # 'monte-carlo' args = [dt]
    def get_result_products(self, method='closed-form', inplace=True, greeks=False):
        if not greeks:
            df = pd.DataFrame(columns=[DICT_RESULTS['price']])
        else:
            df = pd.DataFrame(columns=list(DICT_RESULTS.values()))

        for p in self._portfolio.get_products():
            if method == 'closed-form':
                args = []
            elif method == 'monte-carlo':
                args = [DT]
            else:
                raise Exception('Error: Pricing method undefined.')
        
            df.loc[len(df.index)] = get_pricing_results(p, method, args, greeks)

        if inplace:
            self._df_out_products = df
        
        return df
    
    def write_result_products(self, filename, sheet_idx=0):
        with xw.Book(filename) as wb:
            # last column of input is the first column of output
            start_col = self._skipleft + self._df_in.shape[1] + 1
            start_row = len(self._skiprows) + 1 # hardcoded
            start_cell = get_column_letter(start_col) + str(start_row) # 'O1'

            sheet = wb.sheets[sheet_idx]
            sheet.range(start_cell).options(index=False).value = self._df_out_products

            wb.save()

    def plot_product(self, filename, asset, S_endp, T_endp, N=100, greeks=False):
        if not greeks:
            tasks = ['price']
        else:
            tasks = list(DICT_RESULTS.keys())
        
        # change N to N+1 to avoid Spot = Strike
        S_range = np.linspace(S_endp[0], S_endp[1], N + 1)

        # T: time to maturity
        dt = asset.dt if hasattr(asset, 'dt') else DT
        T_first = np.ceil(T_endp[1] / dt)
        T_last = np.floor(T_endp[0] / dt)
        T_num = T_first - T_last + 1 # max # of observations
        T_size = min(T_num, N)
        T_range = dt * np.linspace(T_last, T_first, T_size, dtype=int)

        self._surface = Surface(S_range, T_range, asset, greeks)

        self._out_dfs = {}
        for task in tasks:
            self._out_dfs[task] = self._surface.table(task)
        
        figs = {}
        
        for task in tasks:
            fig, ax = plt.subplots(
                nrows=1, ncols=1, figsize=(PAPER_SIZE[0], PAPER_SIZE[0]),
                subplot_kw = {"projection": "3d"}
            )
            self._surface.plot(ax, task)
            plt.savefig(f'{task}.png')
            figs[task] = fig

        out_sheetnames = [s[:-4] for s in list(DICT_RESULTS.values())]

        with xw.Book(filename) as wb:
            sheetnames = [s.name for s in wb.sheets]
            
            for task in tasks:
                sheet_name = DICT_RESULTS[task][:-4]
                if sheet_name not in sheetnames:
                    wb.sheets.add(sheet_name)
                sheet = wb.sheets[sheet_name]

                sheet.range('B2').value = self._out_dfs[task].values
                sheet.range('B1').value = self._out_dfs[task].columns.to_list()
                sheet.range('A2').options(transpose=True).value = self._out_dfs[task].index.to_list()

                sheet.pictures.add(figs[task], name=task, update=True)
                plt.close(fig)
            
            wb.save()
        
    def plot_portfolio():
        pass

    # n_cols: # of subplots per row (# of cols)
    # figsize[0]: width of A4
    def plot(self, filename, S_endp, T_endp, N=100, greeks=False):
        products = self.get_products()
        if price_only:
            tasks = ['price']
        else:
            tasks = list(DICT_RESULTS.keys())

        n_rows = len(products) * len(tasks) // 2

        fig, axs = plt.subplots(
            nrows = n_rows, ncols = n_cols,
            figsize = (PAPER_SIZE[0], PAPER_SIZE[0]/2 * n_rows),
            layout = 'constrained', subplot_kw = {"projection": "3d"}
        )
        fig.set_constrained_layout_pads(w_pad = 0.2, h_pad = 0.2)

        plot_id = 0
        plot_row = 0
        plot_col = 0
        
        # change N to N+1 to avoid Spot = Strike
        S_range = np.linspace(S_endp[0], S_endp[1], N + 1)

        # T: time to maturity
        if hasattr(products[0], '_dt'): # discretely monitored
            dt = products[0].dt
            T_first = np.ceil(T_endp[1] / dt)
            T_last = np.floor(T_endp[0] / dt)
            T_num = T_first - T_last + 1 # max # of observations
            T_size = min(T_num, N)
            T_range = dt * np.linspace(T_last, T_first, T_size, dtype=int)
        else:
            T_range = np.linspace(T_endp[0], T_endp[1], N)

        for p in products:
            surface = Surface(S_range, T_range, p)
            surface.generate()

            for task in list(DICT_RESULTS.keys()):
                surface.plot(axs[plot_row][plot_col], task)
                
                plot_id += 1
                plot_row, plot_col = divmod(plot_id, 2)

        plt.savefig(out_filename)
        plt.close(fig)
    
    def write_results(self, out_filename, sheetname):
        # df_print = self._df_out.map(lambda x: '{:.2%}'.format(x))
        # with pd.ExcelWriter(out_filename, mode='a', if_sheet_exists='overlay') as writer:
        #     df_print.to_excel(
        #         writer, sheet_name = sheetname,
        #         header = True, index = False,
        #         startrow = self._skiptop,
        #         startcol = self._skipleft + self._df_in.shape[1]
        #     )

        self._out_filename = out_filename
        if self._in_filename is None:
            raise Exception('Error: Input file undefined.')

        wb = load_workbook(self._in_filename)
        ws = wb[sheetname]

        # 1-based
        start_row = self._skiptop + 1
        start_col = self._skipleft + self._df_in.shape[1] + 1

        for j in range(self._df_out.shape[1]):
            cell = ws.cell(row = start_row, column = start_col + j)
            cell.value = self._df_out.columns[j]
            cell.fill = PatternFill(fill_type='solid', fgColor='1F4E78')
            cell.font = Font(b=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for i in range(self._df_out.shape[0]):
            for j in range(self._df_out.shape[1]):
                cell = ws.cell(row = start_row + 1 + i, column = start_col + j)
                cell.value = self._df_out.iloc[i, j]
                cell.number_format = '0.00%'
                cell.fill = PatternFill(fill_type='solid', fgColor='FFFF00')
                cell.font = Font(color='FF0000')
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        wb.save(out_filename)

    def write_image(self, img_filename, xlsx_filename, sheetname):
        if self._out_filename is None:
            raise Exception('Error: Output file with data is undefined.')
        
        wb = load_workbook(self._out_filename)
        ws = wb[sheetname]
        img = Image(img_filename)

        spacing = 2

        # 1-based
        img_row = self._skiptop + self._df_in.shape[0] + 1 + spacing
        img_col = self._skipleft + 1

        img.anchor = f'{get_column_letter(img_col)}{img_row + 1}'
        ws.add_image(img)
        wb.save(xlsx_filename)
    